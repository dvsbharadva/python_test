import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook("transactions.xlsx")
sheet = wb["Sheet1"]

print(sheet["a1"].value)
cell = sheet.cell(1, 3)
print(cell.value)
print(f"Total rows in sheet are: {sheet.max_row}")

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    print(cell.value)

# def add_xl_automation_with_chart(filename):

for row in range(2, sheet.max_row + 1):
    cell_value = sheet.cell(row, 3)
    discount_price = cell_value.value * 0.8
    discount_cell = sheet.cell(row, 4)
    discount_cell.value = discount_price

values = Reference(
                min_row=2,
                max_row=sheet.max_row,
                min_col=4,
                max_col=4
            )
chart = BarChart()
chart.add_data(values)
# sheet.addChart(chart, "e2")

wb.save("transactions.xlsx")


# add_xl_automation_with_chart("transactions.xlsx")